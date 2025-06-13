from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from .models import UserInfo
from django.utils import timezone
from storages.backends.s3boto3 import S3Boto3Storage


# Excel export action
@admin.action(description='Export Selected to Excel')
def export_to_excel(modeladmin, request, queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "User Info"
    
    # Header row with all fields
    headers = [
        'ID',
        'Full Name',
        'Email Address',
        'WhatsApp Number',
        'Location',
        'Portfolio URL',
        'Experience Description',
        'Experience Level',
        'Primary Editing Software',
        'Specialization',
        'Expected Price',
        'Availability',
        'Why Join Editopia Studio',
        'Resume/CV',
        'Additional Information',
        'Submitted At'
    ]
    sheet.append(headers)
    
    # Data rows with all fields
    for user in queryset:
        sheet.append([
            user.id,
            user.full_name,
            user.email,
            user.whatsapp_number,
            user.location,
            user.portfolio_link,  # This will be converted to hyperlink below
            user.years_of_experience,
            user.experience_level,
            user.software_tools_used,
            user.best_video_type,
            user.expected_price,
            user.weekly_availability,
            user.reason_to_join,
            user.resume.url if user.resume else '',  # This will be converted to hyperlink below
            user.additional_info or '',  # Handle null values
            timezone.localtime(user.created_at).strftime("%Y-%m-%d %I:%M %p")

        ])
    
    # Convert URLs to clickable hyperlinks
    for row_num, user in enumerate(queryset, start=2):  # start=2 because row 1 is headers
        # Portfolio URL hyperlink (column F = 6)
        if user.portfolio_link:
            portfolio_cell = sheet.cell(row=row_num, column=6)
            portfolio_cell.hyperlink = user.portfolio_link
            portfolio_cell.value = user.portfolio_link
            portfolio_cell.style = "Hyperlink"
        
        # Resume hyperlink (column N = 14)
        if user.resume:
            resume_cell = sheet.cell(row=row_num, column=14)
            # Build full URL for resume
            resume_url = request.build_absolute_uri(user.resume.url)
            resume_cell.hyperlink = resume_url
            resume_cell.value = f"Download {user.resume.name.split('/')[-1]}"  # Show just filename
            resume_cell.style = "Hyperlink"
    
    # Auto-adjust column widths for better readability
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Style the header row
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="user_info_complete.xlsx"'
    workbook.save(response)
    return response

# Admin config (updated to show more fields in list view if desired)
@admin.register(UserInfo)
class UserInfoAdmin(admin.ModelAdmin):
    list_display = ('id', 'full_name', 'email', 'experience_level', 'created_at')
    list_filter = ('created_at', 'experience_level', 'best_video_type')
    search_fields = ('full_name', 'email', 'location')
    ordering = ('-id',)
    actions = [export_to_excel]
   

def get_signed_resume_url(resume_field):
    if not resume_field:
        return None

    storage = S3Boto3Storage()
    return storage.url(resume_field.name)  # returns a signed URL with expiry
